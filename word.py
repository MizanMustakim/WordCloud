from wordcloud import WordCloud
import matplotlib.pyplot as plt
from PIL import Image
import numpy as np
import openpyxl

# store the text
wb = openpyxl.load_workbook("Book1.xlsx")  # loading the workbook where I will get the names.
sheet = wb["Sheet1"]
with open("1.txt", "w", encoding="utf-8") as file:  # the path directory of text file where the names will be stored
    # from xl file.
    for row in range(1, sheet.max_row + 1):
        for column in range(2, 3):
            cell = sheet.cell(row, column)
            file.write(f"{cell.value}\n")


# processing word cloud
def cloud(text):
    mask = np.array(Image.open("bit.png"))
    wc = WordCloud(background_color='black',
                   min_font_size=12,
                   max_font_size=35,
                   contour_width=.5,
                   contour_color="white",
                   mask=mask)
    wc.generate(text)
    plt.figure(figsize=[10, 10])
    plt.imshow(wc, interpolation='bilinear')
    plt.axis("off")
    plt.show()


get_text = open("1.txt", "r", encoding="utf-8").read()
cloud(get_text)  # calling the function
